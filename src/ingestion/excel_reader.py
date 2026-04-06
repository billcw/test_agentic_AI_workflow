"""
excel_reader.py - Excel spreadsheet text extraction.
"""

from pathlib import Path
from openpyxl import load_workbook


def read_excel(file_path: str | Path) -> list[dict]:
    """
    Extract text from an Excel file (.xlsx).
    Each sheet becomes one entry, with rows formatted as text.

    Returns a list of dicts, one per sheet:
    [
        {
            "page": 1,
            "text": "Sheet: Sheet1\nRow 1: val1 | val2...",
            "method": "excel",
            "source": "filename.xlsx"
        },
        ...
    ]
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(str(file_path), data_only=True)
    sheets = []

    for sheet_num, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        rows = []

        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            values = [str(cell) for cell in row if cell is not None]
            if values:
                rows.append(" | ".join(values))

        if rows:
            sheets.append({
                "page": sheet_num + 1,
                "text": f"Sheet: {sheet_name}\n" + "\n".join(rows),
                "method": "excel",
                "source": file_path.name
            })

    return sheets
